#!/usr/bin/env python3
"""Merge ZoomInfo phone-lookup batch results into both master artifacts."""
import json
from pathlib import Path

import gspread
import openpyxl

DATA = Path(__file__).parent / "data"
MASTER = DATA / "AZ_targets_enriched_master.xlsx"


def merge(output_files, log_label):
    results = []
    for f in output_files:
        p = DATA / f
        if p.exists():
            results.extend(json.load(open(p)))
    by_id = {r["id"]: r for r in results}
    print(f"[{log_label}] total usable results: {len(by_id)}")

    wb = openpyxl.load_workbook(MASTER)
    ws = wb["Enriched Master"]
    header = [c.value for c in ws[1]]
    col = {n: i for i, n in enumerate(header)}

    updated = []
    for row in ws.iter_rows(min_row=2):
        if not row[1].value or row[0].value not in by_id:
            continue
        r = by_id[row[0].value]
        if r.get("mobile_phone"):
            row[col["Owner Cell"]].value = r["mobile_phone"]
            row[col["Cell Source"]].value = "ZoomInfo"
            if "donotcall" in (r.get("match_notes") or "").lower().replace(" ", "").replace("_", ""):
                row[col["Cell DNC Status"]].value = "DNC -- do not call"
        elif r.get("office_phone") and not row[col["Office Phone"]].value:
            row[col["Office Phone"]].value = r["office_phone"]
        if r.get("email") and not row[col["Owner Email"]].value:
            row[col["Owner Email"]].value = r["email"]
        if r.get("match_notes"):
            existing = row[col["Notes"]].value or ""
            note = f"ZoomInfo: {r['match_notes']}"
            row[col["Notes"]].value = f"{existing}; {note}" if existing else note
        updated.append(row[0].value)

    wb.save(MASTER)
    print(f"[{log_label}] local xlsx: updated {len(updated)} rows")

    gc = gspread.service_account(filename=str(Path.home() / "dev/hvac-lead-sourcing/service_account.json"))
    sh = gc.open_by_key("1bbOBPow3M9a4dgEtQ2wodtdlkXy_fcAiyexKz7yfsQ0")
    gsheet_ws = sh.worksheet("Enriched Master")
    all_values = gsheet_ws.get_all_values()
    gheader = all_values[0]
    gcol = {n: i for i, n in enumerate(gheader)}
    id_to_row = {r[0]: i + 2 for i, r in enumerate(all_values[1:]) if r and r[0]}

    wb2 = openpyxl.load_workbook(MASTER)
    ws2 = wb2["Enriched Master"]
    cell_updates = []
    for row in ws2.iter_rows(min_row=2):
        if row[0].value in updated and row[0].value in id_to_row:
            grow = id_to_row[row[0].value]
            cell_updates.append(gspread.Cell(grow, gcol["Owner Cell"] + 1, row[col["Owner Cell"]].value))
            cell_updates.append(gspread.Cell(grow, gcol["Office Phone"] + 1, row[col["Office Phone"]].value))
            cell_updates.append(gspread.Cell(grow, gcol["Owner Email"] + 1, row[col["Owner Email"]].value))
            cell_updates.append(gspread.Cell(grow, gcol["Cell Source"] + 1, row[col["Cell Source"]].value))
    if cell_updates:
        gsheet_ws.update_cells(cell_updates, value_input_option="RAW")
    print(f"[{log_label}] live Google Sheet: updated {len(updated)} rows ({len(cell_updates)} cells)")


if __name__ == "__main__":
    import sys
    files = sys.argv[1:]
    merge(files, "merge")

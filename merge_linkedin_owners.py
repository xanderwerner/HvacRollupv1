#!/usr/bin/env python3
"""Merge the 10 LinkedIn/web research batches into both master artifacts.

Exclusions applied (companies found but judged not real acquisition targets):
- CED / Consolidated Electrical Distributors branches -- national wholesale distributor,
  not an independent contractor.
- Strada Services -- large multi-state company, wrong target type, confirmed twice now
  (once via Apollo, once via this LinkedIn pass).
- Automotive A/C shops -- wrong industry (car AC, not building HVAC).

Flags applied (owner written in, but Data Check notes a real concern):
- Deceased owner, licensing issues, acquired-by-competitor, or duplicate-listing risk.
"""
import json
import re
from pathlib import Path

import gspread
import openpyxl

DATA = Path(__file__).parent / "data"
MASTER = DATA / "AZ_targets_enriched_master.xlsx"

EXCLUDE_IDS = {
    # CED / national distributor branches
    "C2810",  # Vegas Electrical Supply Co (CED branch, batch 10)
}
EXCLUDE_NAME_SUBSTR = ["consolidated electrical distributors", "strada services",
                        "magics auto ac", "auto air conditioning repair shop"]

FLAG_NOTES = {
    "C2811": "Owner Tristan Scudder died June 1, 2024 (obituary confirmed) -- business status post-death unclear, verify before outreach",
    "C2815": "Related to C2811 (Scudder family) -- both plausible owners (father Joel d.2020, son Tristan d.2024) are deceased, no living owner found",
    "C821": "Contractor license shows suspended status as of 12/30/2023 -- verify current standing before outreach",
    "C855": "Owner matches a 2020 AZ ROC news item re: contracting without a license -- light check before outreach",
    "C683": "Founder reportedly joined ResiXperts (a competing HVAC roll-up) June 2025 -- ownership structure post-deal unclear, re-verify",
    "C864": "Appears to be a multi-state marketing/lead-gen brand (AZ/CA/FL/NV/TX), not a single-owner business -- may not fit target profile",
    "C874": "Wise Electric domain (wisesolarsolutionsinc.com) suggests solar work -- check against solar-exclusion rule before outreach",
}

records = []
for i in range(1, 11):
    records.extend(json.load(open(DATA / f"li_output{i}.json")))

confirmed = []
for r in records:
    if not r.get("owner_name"):
        continue
    if r["id"] in EXCLUDE_IDS:
        continue
    name_lower = r.get("company_name", "").lower()
    if any(s in name_lower for s in EXCLUDE_NAME_SUBSTR):
        continue
    confirmed.append(r)

print(f"total records: {len(records)}, with owner: {sum(1 for r in records if r.get('owner_name'))}, "
      f"after exclusions: {len(confirmed)}")

by_id = {r["id"]: r for r in confirmed}

wb = openpyxl.load_workbook(MASTER)
ws = wb["Enriched Master"]
header = [cell.value for cell in ws[1]]
col = {name: i for i, name in enumerate(header)}

updated = []
for row in ws.iter_rows(min_row=2):
    if not row[1].value:
        continue
    rid = row[0].value
    if rid not in by_id:
        continue
    r = by_id[rid]
    row[col["Owner Name"]].value = r["owner_name"]
    row[col["Owner Title"]].value = r.get("owner_title", "")
    if r.get("owner_linkedin_url"):
        row[col["Owner LinkedIn"]].value = r["owner_linkedin_url"]
    row[col["Owner Found Via"]].value = f"Web/LinkedIn research ({r.get('source', 'web')})"
    existing_notes = row[col["Notes"]].value or ""
    note_add = f"Owner research: {r.get('confidence_notes', '')}"
    row[col["Notes"]].value = f"{existing_notes}; {note_add}" if existing_notes else note_add
    if rid in FLAG_NOTES:
        existing_dc = row[col["Data Check"]].value or ""
        row[col["Data Check"]].value = f"{existing_dc}; {FLAG_NOTES[rid]}" if existing_dc else FLAG_NOTES[rid]
    updated.append(rid)

wb.save(MASTER)
print(f"local xlsx: updated {len(updated)} rows")

gc = gspread.service_account(filename=str(Path.home() / "dev/hvac-lead-sourcing/service_account.json"))
sh = gc.open_by_key("1bbOBPow3M9a4dgEtQ2wodtdlkXy_fcAiyexKz7yfsQ0")
gsheet_ws = sh.worksheet("Enriched Master")
all_values = gsheet_ws.get_all_values()
gheader = all_values[0]
gcol = {name: i for i, name in enumerate(gheader)}
id_to_row = {r[0]: i + 2 for i, r in enumerate(all_values[1:]) if r and r[0]}

cell_updates = []
for rid in updated:
    if rid not in id_to_row:
        continue
    r = by_id[rid]
    grow = id_to_row[rid]
    cell_updates.append(gspread.Cell(grow, gcol["Owner Name"] + 1, r["owner_name"]))
    cell_updates.append(gspread.Cell(grow, gcol["Owner Title"] + 1, r.get("owner_title", "")))
    if r.get("owner_linkedin_url"):
        cell_updates.append(gspread.Cell(grow, gcol["Owner LinkedIn"] + 1, r["owner_linkedin_url"]))
    cell_updates.append(gspread.Cell(grow, gcol["Owner Found Via"] + 1, f"Web/LinkedIn research ({r.get('source', 'web')})"))
    if rid in FLAG_NOTES:
        existing_dc = all_values[grow - 1][gcol["Data Check"]] if grow - 1 < len(all_values) else ""
        new_dc = f"{existing_dc}; {FLAG_NOTES[rid]}" if existing_dc else FLAG_NOTES[rid]
        cell_updates.append(gspread.Cell(grow, gcol["Data Check"] + 1, new_dc))

gsheet_ws.update_cells(cell_updates, value_input_option="RAW")
print(f"live Google Sheet: updated {len(updated)} rows ({len(cell_updates)} cells)")

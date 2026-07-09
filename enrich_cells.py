#!/usr/bin/env python3
"""Batch-reveal owner mobile numbers via Apollo people/match + webhook_result polling.

Targets: master rows with Owner Name but no Owner Cell and a domain or email.
Priority: ICP hotlist tiers B/B2 first, then by Google review count.
Progress is appended to reveals.jsonl (resumable — already-processed IDs are skipped).

Usage: python3 enrich_cells.py [limit]
"""
import json
import re
import sys
import time
from pathlib import Path

import ssl

import certifi
import openpyxl
import urllib.request

SSL_CTX = ssl.create_default_context(cafile=certifi.where())

BASE = Path(__file__).parent / "data"
MASTER = BASE / "AZ_targets_enriched_master.xlsx"
HOTLIST = BASE / "ICP_hotlist_v2.xlsx"
LOG = BASE / "reveals.jsonl"
KEY = next(
    line.split("=", 1)[1].strip()
    for line in open(Path.home() / "dev/hvac-lead-sourcing/.env")
    if line.startswith("APOLLO_API_KEY=")
)
WEBHOOK = "https://example.com/apollo-callback"  # delivery fails, result still pollable


def api(path, payload=None):
    url = f"https://api.apollo.io/api/v1/{path}"
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode() if payload is not None else None,
        headers={"X-Api-Key": KEY, "Content-Type": "application/json"},
        method="POST" if payload is not None else "GET",
    )
    last = None
    for attempt in range(5):
        try:
            with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as r:
                return json.loads(r.read())
        except urllib.error.HTTPError as e:
            if e.code == 429:
                last = "429"
                time.sleep(45 * (attempt + 1))
                continue
            body = e.read().decode()[:300]
            return {"_error": e.code, "_body": body}
        except Exception as e:  # network blips
            time.sleep(5 * (attempt + 1))
            last = str(e)
    return {"_error": "retries_exhausted", "_body": last}


def load_targets(limit):
    wb = openpyxl.load_workbook(MASTER)
    rows = list(wb["Enriched Master"].iter_rows(values_only=True))[1:]
    hot = openpyxl.load_workbook(HOTLIST)
    tier_by_id = {
        r[1]: str(r[0])[:2].strip()
        for r in list(hot["ICP Hot List v2"].iter_rows(values_only=True))[1:]
        if r[0] and r[1]
    }
    # A row is done if: match failed (no person), or a poll landed a real result
    # (phones, or a clean empty). Poll errors (404/timeout = throttle casualties)
    # and matched-but-never-polled rows get requeued.
    done = set()
    if LOG.exists():
        for line in open(LOG):
            try:
                r = json.loads(line)
            except Exception:
                continue
            if not r.get("poll"):
                if not r.get("matched"):
                    done.add(r["row_id"])
            elif r.get("phones") or not r.get("error"):
                done.add(r["row_id"])
    targets = []
    for r in rows:
        rid, name, owner, cell, email, domain = r[0], r[1], r[5], r[7], r[9], r[22]
        if not owner or cell or rid in done:
            continue
        if not domain and not email:
            continue
        tier = tier_by_id.get(rid, "")
        prio = 0 if tier in ("B", "B2") else 1
        reviews = r[27] or 0
        targets.append((prio, -reviews, rid, name, str(owner).strip(), email, domain, tier))
    targets.sort()
    return targets[:limit]


WAVE = 15
POLL_EVERY = 15
POLL_MAX_SECS = 480


def extract_phones(res):
    wr = res.get("webhook_result") or {}
    phones = []
    for p in wr.get("people", []):
        for ph in p.get("phone_numbers", []):
            phones.append(
                {
                    "number": ph.get("sanitized_number"),
                    "type": ph.get("type_cd"),
                    "status": ph.get("status_cd"),
                    "confidence": ph.get("confidence_cd"),
                    "dnc": ph.get("dnc_status_cd"),
                }
            )
    return phones, wr.get("credits_consumed")


def revealed_person_ids():
    """person_ids that already have a phone result logged — never re-reveal (8 credits each)."""
    ids = set()
    if not LOG.exists():
        return ids
    match_by_row = {}
    for line in open(LOG):
        try:
            r = json.loads(line)
        except Exception:
            continue
        if not r.get("poll") and r.get("person_id"):
            match_by_row[r["row_id"]] = r["person_id"]
        if r.get("poll") and r.get("phones"):
            pid = match_by_row.get(r["row_id"])
            if pid:
                ids.add(pid)
    return ids


def main():
    limit = int(sys.argv[1]) if len(sys.argv) > 1 else 10_000
    targets = load_targets(limit)
    already = revealed_person_ids()
    print(f"targets this run: {len(targets)}", flush=True)
    stats = {"match": 0, "mobile": 0, "other": 0, "none": 0}
    log = open(LOG, "a")
    for w in range(0, len(targets), WAVE):
        wave = targets[w : w + WAVE]
        pending = []
        for prio, negrev, rid, company, owner, email, domain, tier in wave:
            payload = {
                "name": owner,
                "organization_name": company,
                "reveal_phone_number": True,
                "webhook_url": WEBHOOK,
            }
            if domain:
                payload["domain"] = str(domain).strip()
            if email:
                payload["email"] = str(email).strip()
            resp = api("people/match", payload)
            person = resp.get("person") or {}
            rec = {
                "row_id": rid,
                "company": company,
                "owner_input": owner,
                "tier": tier,
                "matched": bool(person),
                "person_id": person.get("id"),
                "person_name": person.get("name"),
                "person_title": person.get("title"),
                "person_email": person.get("email"),
                "linkedin": person.get("linkedin_url"),
                "phone_request_id": resp.get("request_id"),
                "error": resp.get("_error"),
                "ts": time.strftime("%Y-%m-%dT%H:%M:%S"),
            }
            log.write(json.dumps(rec) + "\n")
            log.flush()
            if rec["matched"]:
                stats["match"] += 1
                if rec["person_id"] in already:
                    log.write(json.dumps({"row_id": rid, "poll": True, "phones": [],
                                          "note": "person already revealed, see other row",
                                          "ts": time.strftime("%Y-%m-%dT%H:%M:%S")}) + "\n")
                elif rec["phone_request_id"] is not None:
                    pending.append({**rec, "submitted": time.time()})
            time.sleep(2.5)

        while pending:
            time.sleep(POLL_EVERY)
            still = []
            for rec in pending:
                res = api(f"webhook_result/{rec['phone_request_id']}")
                if res.get("webhook_result"):
                    phones, credits = extract_phones(res)
                    log.write(json.dumps({"row_id": rec["row_id"], "poll": True, "phones": phones,
                                          "credits": credits,
                                          "ts": time.strftime("%Y-%m-%dT%H:%M:%S")}) + "\n")
                    log.flush()
                    if rec["person_id"]:
                        already.add(rec["person_id"])
                    mob = [p for p in phones if p["type"] == "mobile"]
                    if mob:
                        stats["mobile"] += 1
                        tag = "MOBILE " + mob[0]["number"]
                    elif phones:
                        stats["other"] += 1
                        tag = "other " + phones[0]["number"]
                    else:
                        stats["none"] += 1
                        tag = "no phone"
                    print(f"{rec['company'][:44]:44s} {tag}", flush=True)
                elif (
                    res.get("_error") == 404 and time.time() - rec["submitted"] > 90
                ) or time.time() - rec["submitted"] > POLL_MAX_SECS:
                    log.write(json.dumps({"row_id": rec["row_id"], "poll": True, "phones": [],
                                          "error": res.get("_error") or "timeout",
                                          "ts": time.strftime("%Y-%m-%dT%H:%M:%S")}) + "\n")
                    log.flush()
                    stats["none"] += 1
                    print(f"{rec['company'][:44]:44s} no result ({res.get('_error') or 'timeout'})", flush=True)
                else:
                    still.append(rec)
                time.sleep(0.4)
            pending = still
        done = min(w + WAVE, len(targets))
        print(f"--- wave done {done}/{len(targets)} | {stats}", flush=True)
        time.sleep(30)  # inter-wave cooldown, stay under Apollo rate limits
    print(f"FINISHED {stats}", flush=True)


if __name__ == "__main__":
    main()

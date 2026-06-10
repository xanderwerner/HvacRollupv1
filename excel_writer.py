"""Writes qualified leads into the Organic List tab of the Excel database.

Design rules (from the v1 spec):
  * Map by HEADER TEXT (row 4) — never hard-code column letters.
  * Only write dropdown-SAFE values so data-validation never breaks.
  * Append after existing rows; dedupe against company names already present.
  * Never touch headers, the example row, or formatting.
"""
import datetime
from typing import Optional

import openpyxl

import config
from models import Lead


def _clean_url(url: str) -> str:
    if not url:
        return ""
    for p in ("https://", "http://"):
        if url.startswith(p):
            url = url[len(p):]
    if url.startswith("www."):
        url = url[4:]
    return url.rstrip("/")


def _row_values(lead: Lead, prospect_no: int, date_str: str) -> dict:
    """Header-text -> cell value. Blank = pending enrichment (v2)."""
    note = (
        f"Auto-sourced via Google Places (v1 free pipeline) {date_str}. "
        f"Buy-box gate: {lead.review_count} reviews, {lead.rating}★. "
        f"PENDING: ROC license/owner, ACC age, owner contact (Apollo/ZoomInfo), motivation."
    )
    if lead.business_phone:
        note += f" Main line: {lead.business_phone}."

    return {
        "Prospect #": f"{prospect_no:03d}",
        "Researched By": "",                 # unclaimed — a human claims it
        "Research Stage": "New",
        "Disqualify Reason": "",
        "Discovery Source": "Google Maps",
        "Ad Platform Seen On": "N/A",
        "Date Added": date_str,
        "Company Name": lead.company_name,
        "Trade Type": lead.trade_type,
        "City": lead.city,
        "Business Address": lead.business_address,
        "Website": _clean_url(lead.website),
        "Years in Business": "",
        "ROC License #": "",
        "ROC License Status": "",
        "ACC Entity Name": "",
        "ACC Formation Date": "",
        "Google Review Count": lead.review_count,
        "Google Rating": lead.rating,
        "Est. # of Trucks": "",
        "Est. # of Employees": "",
        "ZoomInfo Rev. Estimate": "",
        "Running Ads": "Unknown",
        "Ad Activity Notes": "",
        "Book of Business Signal": "",
        "Owner Name": "",
        "Owner Title": "",
        "Owner Direct Mobile": "",
        "Owner Email": "",
        "Owner LinkedIn": "",
        "Owner Tenure": "",
        "Owner Age Estimate": "Unknown",
        "ZoomInfo Intent Signal": "Not checked",
        "Ever Listed for Sale": "Unknown",
        "GM / Ops Job Posting": "Unknown",
        "Succession Gap": "Unknown",
        "Owner Headspace Notes": "",
        "Motivation Score": "",
        "ROC Checked": "No",
        "ACC Checked": "No",
        "Google Checked": "Yes",
        "ZoomInfo Checked": "No",
        "LinkedIn Checked": "No",
        "BBB / Reviews Checked": "No",
        "PE / Roll-up Check": "Unknown",
        "Buy Box Fit": lead.buy_box_fit,
        "Priority Tier": lead.priority_tier,
        "Ready to Hand Off": "Not yet",
        "Notes": note,
    }


def write_leads(leads: list[Lead], limit: Optional[int] = None) -> list[str]:
    """Append survivor leads to the Organic List. Returns names written."""
    wb = openpyxl.load_workbook(config.EXCEL_PATH)
    if config.EXCEL_SHEET not in wb.sheetnames:
        raise RuntimeError(f"Sheet '{config.EXCEL_SHEET}' not found")
    ws = wb[config.EXCEL_SHEET]

    # header text -> column index
    headers = {
        c.value: c.column
        for c in ws[config.HEADER_ROW]
        if c.value is not None
    }
    col_company = headers["Company Name"]
    col_prospect = headers["Prospect #"]

    # existing state: company names already present + max prospect number
    existing_names, max_no, write_row = set(), 0, None
    r = config.FIRST_DATA_ROW
    while True:
        name = ws.cell(row=r, column=col_company).value
        if name in (None, ""):
            write_row = write_row or r
            # keep scanning a little to be safe, but first empty is our start
            break
        existing_names.add(str(name).strip().lower())
        pv = ws.cell(row=r, column=col_prospect).value
        try:
            max_no = max(max_no, int(str(pv).strip()))
        except (TypeError, ValueError):
            pass
        r += 1
    write_row = write_row or r

    date_str = datetime.date.today().strftime("%m/%d/%y")
    written, no = [], max_no
    for lead in leads:
        if not lead.survived_filter:
            continue
        if lead.company_name.strip().lower() in existing_names:
            continue
        no += 1
        for header, value in _row_values(lead, no, date_str).items():
            if header in headers and value != "":
                ws.cell(row=write_row, column=headers[header], value=value)
        existing_names.add(lead.company_name.strip().lower())
        written.append(lead.company_name)
        write_row += 1
        if limit and len(written) >= limit:
            break

    wb.save(config.EXCEL_PATH)
    return written

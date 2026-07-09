#!/usr/bin/env python3
"""Deduplicate multi-location companies in the master (post ROC-revert-fix).

Merge groups are curated from manual verification (license-number agreement is
the strongest signal; brand-name + domain agreement is the fallback for rows
missing license data). Same-owner-but-genuinely-different-company situations
are NOT merged -- they're cross-referenced in Notes so nobody re-buys a phone
number we already have for that person.
"""
import json
import openpyxl

DATA = "/Users/xanderwerner/Desktop/hvac/data/"

# Each group: canonical row_id first, then the duplicate row_ids to fold in and delete.
# Basis noted per group for the audit trail.
MERGE_GROUPS = [
    # --- confirmed via matching ROC license number ---
    (["C219v2"], "placeholder"),  # unused, real list below
]

MERGE_GROUPS = [
    {"keep": None, "ids": ["C_AC_REPAIR_PLACEHOLDER"], "basis": "placeholder"},
]

# Real curated groups keyed by company-name substrings (resolved to row_ids at runtime)
NAME_BASED_GROUPS = [
    {"names": ["AC Repair Near Me", "AC Repair Near Me Of Scottsdale", "AC Repair Near Me of Peoria",
               "AC Repair Near Me of Phoenix", "AC Repair Near Me of Tempe", "AC Repair Near Me of Goodyear",
               "AC Repair Near Me of Gilbert"],
     "basis": "same license 337558"},
    {"names": ["Scottsdale Plumbing Co", "Scottsdale Air Heating & Cooling", "Scottsdale Heating and Cooling"],
     "basis": "same license 268561"},
    {"names": ["America’s Home Services Of Chandler", "America’s Home Services Of Gilbert"],
     "basis": "same license 345894"},
    {"names": ["Cool Breeze Comfort Solutions", "Cool Breeze AC Repair Surprise AZ"], "basis": "same license 289247"},
    {"names": ["Precision Air & Plumbing", "Precision Plumbing"], "basis": "same license 126405"},
    {"names": ["Phend Plumbing & Rooter LLC", "Phend Plumbing and Rooter LLC"], "basis": "same license 288046"},
    {"names": ["Just Air LLC", "Just Plumbing"], "basis": "same license 221611"},
    {"names": ["North Mechanical Heating & Air Conditioning", "North Mechanical Heating and Cooling"],
     "basis": "same license 196159"},
    {"names": ["Cactus Plumbing And Air", "Cactus Plumbing & Air"], "basis": "same license 336756"},
    {"names": ["My Plumber Plumbing Services LLC", "P&R Plumbing Services And Air LLC"],
     "basis": "same license 337614"},
    {"names": ["Arizona Cooling and Heating Specialists LLC", "Cooling Specialists"], "basis": "same license 345033"},
    {"names": ["All Season Plumbing And Air", "All Season Heating & Air Conditioning"], "basis": "same license 348428"},
    {"names": ["Platinum Air Heating & Cooling, LLC", "Platinum Services Plumbing"], "basis": "same license 355439"},
    {"names": ["Roto-Rooter Plumbing and Drain Service", "Roto-Rooter Plumbing Drain Services"],
     "basis": "same license 159873"},
    {"names": ["Action Air Conditioning", "Action Plumbing"], "basis": "same license 339126"},
    {"names": ["Oro Valley Plumbing LLC", "Oro Valley Heating and Cooling LLC"], "basis": "same license 285210"},
    {"names": ["Queen Creek Air Conditioning", "Queen Creek Plumbing Services"], "basis": "same license 286725"},
    {"names": ["NextGen Plumbing LLC", "Next Plumbing & Heating"], "basis": "same license 355868"},
    {"names": ["Reliable Plumbing", "Reliable Plumbing & Heating"], "basis": "same license 66053"},
    {"names": ["Jobe & Sons Plumbing", "Jobe & Sons Plumbing- Northwest"], "basis": "same license 336499"},
    {"names": ["Prescott Valley Heating and Cooling LLC", "Prescott Valley Heating & Air Conditioning"],
     "basis": "same license 333250"},
    {"names": ["Felix Appliance Heating & AC Repair", "Felix Appliance Heating & Air in Casa Grande"],
     "basis": "same license 322603"},
    {"names": ["AirTime Cooling and Heating", "Air Time"], "basis": "same license 351313"},
    # --- confirmed via identical domain ---
    {"names": ["HVAC Near Me", "HVAC Near Me Repair and Installation of Peoria",
               "HVAC Near Me Repair and Installation of Scottsdale", "HVAC Near Me Repair And Installation Of Tempe",
               "HVAC Near Me Repair and Installation of Goodyear"],
     "basis": "same domain hvacnearme.today"},
    # --- brand-name pattern, no owner conflict, no contradicting license ---
    {"names": ["AZ Home Services Group AC Repair & Plumbing Services",
               "AZ Home Services Group of Queen Creek",
               "AZ Home Services Group AC Repair & Plumbing Services of Mesa",
               "AZ Home Services Group AC Repair & Plumbing Services of Scottsdale",
               "AZ Home Services Group AC Repair & Plumbing Services of Gilbert",
               "AZ Home Services Group AC Repair & Plumbing Services of Goodyear"],
     "basis": "identical brand name, city-suffix pattern (owner field was a placeholder artifact)"},
    {"names": ["Plumbers Near Me - Water Heater & Plumbing Services",
               "Plumbers Near Me, Water Heater, and Plumbing Services of Goodyear",
               "Plumbers Near Me, Water Heater, and Plumbing Services of Peoria",
               "Plumbers Near Me – Water Heater & Plumbing Services of Tempe"],
     "basis": "identical brand name, city-suffix pattern (owner field was a placeholder artifact)"},
    {"names": ["Temperature Control, Inc. A/C-Heating & Plumbing - Tucson, AZ",
               "Temperature Control, Inc. A/C-Heating & Plumbing - Oro Valley, AZ"],
     "basis": "same owner (Vance Macleod), identical brand, city-suffix"},
    {"names": ["Rite Way Heating, Cooling & Plumbing", "Rite Way Heating, Cooling & Plumbing: Sierra Vista"],
     "basis": "same owner (Richard Walter), identical brand, city-suffix"},
    {"names": ["Goettl Air Conditioning and Plumbing - Phoenix AZ", "Goettl Air Conditioning and Plumbing - Tucson AZ"],
     "basis": "same owner (Jake Gress), identical brand, city-suffix -- FLAG: large regional brand, verify buy-box fit"},
    {"names": ["ACE Home Services (formerly Yavapai Plumbing & Heating)", "ACE Home Services (formerly AZ Best Plumbing)"],
     "basis": "same owner (Jay Hazlewood), identical current brand name"},
]

# Same owner, genuinely DIFFERENT companies (different license #s) -- cross-reference only, do not merge
CROSS_REF_ONLY = [
    (["Ambient Edge Cooling, Heating & Plumbing", "Day & Night Air Conditioning, Heating, & Plumbing",
      "Plumbing by Jake"], "Steven Ray Lewis"),
    (["Felix Appliance Heating & Air"], "Robert Felix",
     ["Felix Appliance Heating & AC Repair / Casa Grande (merged, same owner)"]),
    (["American Conditioned Air", "Capstone Cooling & Heating LLC"], "Michael Wolters"),
    (["Zest Plumbing & Drain", "Affordable Plumbing, Rooter and Water Heaters"], "Ronald A Degner"),
    (["Reliance Heating & Air Conditioning", "Prescott Heating & Cooling"], "Dalen Thomas Blumentritt"),
    (["Rescue One Air Cooling, Heating & Plumbing"], "Daniel Joseph Santoro",
     ["AirTime Cooling and Heating / Air Time (merged, same owner)"]),
]


def find_row_by_name(rows, name):
    for r in rows:
        if r[1] and str(r[1]).strip() == name.strip():
            return r
    return None


def merge_group(rows_dict, names, basis):
    found = [(n, rows_dict.get(n)) for n in names]
    missing = [n for n, r in found if r is None]
    present = [(n, r) for n, r in found if r is not None]
    if len(present) < 2:
        return None, missing
    # canonical: prefer the one with a domain, then most non-empty fields, then shortest name (base brand)
    def score(item):
        n, r = item
        return (bool(r[22]), sum(1 for v in r if v not in (None, "")), -len(n))
    present.sort(key=score, reverse=True)
    keep_name, keep_row = present[0]
    others = present[1:]
    return {
        "keep_id": keep_row[0], "keep_name": keep_name,
        "other_ids": [r[0] for _, r in others],
        "other_names": [n for n, _ in others],
        "other_cities": [r[3] for _, r in others],
        "basis": basis,
    }, missing


def main():
    wb = openpyxl.load_workbook(DATA + "AZ_targets_enriched_master.xlsx")
    ws = wb["Enriched Master"]
    all_rows = list(ws.iter_rows(values_only=True))[1:]
    by_name = {r[1]: r for r in all_rows if r[1]}

    plans = []
    all_missing = []
    for g in NAME_BASED_GROUPS:
        plan, missing = merge_group(by_name, g["names"], g["basis"])
        if plan:
            plans.append(plan)
        all_missing.extend(missing)

    if all_missing:
        print("WARNING - names not found in master (skipped):", all_missing)

    print(f"\n{len(plans)} merge groups resolved, folding {sum(len(p['other_ids']) for p in plans)} duplicate rows\n")
    for p in plans:
        print(f"KEEP {p['keep_id']} {p['keep_name'][:40]:40} <- fold {p['other_ids']} ({p['basis']})")

    ids_to_delete = set()
    keep_row_by_id = {r[0]: r for r in all_rows}
    updates = {}  # keep_id -> dict of field updates
    for p in plans:
        keep_id = p["keep_id"]
        other_rows = [keep_row_by_id[oid] for oid in p["other_ids"]]
        keep_row = keep_row_by_id[keep_id]
        upd = {}
        # union Other Locations (col 23, 0-based idx23)
        loc_names = [f"{n} ({c})" for n, c in zip(p["other_names"], p["other_cities"])]
        existing_loc = keep_row[23] or ""
        upd["other_locations"] = (str(existing_loc) + "; " if existing_loc else "") + "; ".join(loc_names)
        # fill any blank field on keep_row from the other rows (first non-empty wins)
        fill_idx = [5,6,7,8,9,10,11,12,13,14,17,18,19,20,21,22,24,25,26,27]  # owner..reviews, skip name/city/state/trade
        for idx in fill_idx:
            if not keep_row[idx]:
                for orow in other_rows:
                    if orow[idx]:
                        upd.setdefault("fill", {})[idx] = orow[idx]
                        break
        upd["note"] = f"MERGED multi-location duplicate(s) 2026-07-09: {', '.join(p['other_names'])} ({p['basis']})"
        updates[keep_id] = upd
        ids_to_delete.update(p["other_ids"])

    # cross-reference notes for same-owner-different-company situations
    cross_ref_notes = {}
    for names, owner, *rest in CROSS_REF_ONLY:
        sibling_note_extra = rest[0] if rest else []
        for i, n in enumerate(names):
            r = by_name.get(n)
            if not r:
                continue
            siblings = [x for x in names if x != n] + sibling_note_extra
            if siblings:
                cross_ref_notes[r[0]] = f"NOTE: owner {owner} also owns: {'; '.join(siblings)} -- do not re-buy this person's phone number for those rows"

    # apply everything
    new_rows_written = 0
    for row in ws.iter_rows(min_row=2):
        rid = row[0].value
        if rid in updates:
            u = updates[rid]
            row[23].value = u["other_locations"]
            for idx, val in u.get("fill", {}).items():
                row[idx].value = val
            prev = row[15].value
            row[15].value = (str(prev) + "; " if prev else "") + u["note"]
            new_rows_written += 1
        if rid in cross_ref_notes:
            prev = row[28].value
            row[28].value = (str(prev) + " | " if prev else "") + cross_ref_notes[rid]

    # delete duplicate rows (bottom-up to keep row indices valid)
    rows_to_delete_idx = []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row[0].value in ids_to_delete:
            rows_to_delete_idx.append(i)
    for i in sorted(rows_to_delete_idx, reverse=True):
        ws.delete_rows(i)

    wb.save(DATA + "AZ_targets_enriched_master.xlsx")
    print(f"\ncanonical rows updated: {new_rows_written}")
    print(f"duplicate rows deleted: {len(rows_to_delete_idx)}")
    print(f"cross-reference notes added: {len(cross_ref_notes)}")

    json.dump(sorted(ids_to_delete), open(DATA + "deleted_row_ids.json", "w"))
    print("\ndeleted IDs saved to data/deleted_row_ids.json (for hotlist/tracker/Super Enrichment cleanup)")


if __name__ == "__main__":
    main()

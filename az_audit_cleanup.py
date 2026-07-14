#!/usr/bin/env python3
"""Full data-quality audit + cleanup pass on the AZ master, per Xander's "double check
everything" request:

1. Remove solar companies (violates the standing no-solar rule -- missed during the
   electrical widening/Places sweep since that predates the rule).
2. Remove CED / Consolidated Electrical Distributors branches (national wholesale
   distributor, not an independent contractor -- not a real acquisition target).
3. Remove Strada Services row itself (already excluded from owner enrichment as a
   wrong-target-type; the row was never actually deleted).
4. Flag (Data Check, not delete) same-domain groups found across the whole list --
   these are either legitimate multi-location franchises OR a domain-scrape mismatch
   from earlier sourcing; can't tell which without per-row verification, so flag rather
   than guess. Same-city pairs get a stronger "likely duplicate" flag; different-city
   pairs get a softer "possible franchise or data mismatch" flag.
"""
import re
from collections import defaultdict
from pathlib import Path

import gspread
import openpyxl

DATA = Path(__file__).parent / "data"
MASTER = DATA / "AZ_targets_enriched_master.xlsx"

wb = openpyxl.load_workbook(MASTER)
ws = wb["Enriched Master"]
header = [cell.value for cell in ws[1]]
col = {name: i for i, name in enumerate(header)}

all_rows = [row for row in ws.iter_rows(min_row=2) if row[1].value]

# ---- 1-3: removals ----
KNOWN_FRANCHISE_DOMAINS = {
    "onehourheatandair.com", "aireserv.com", "mrrooter.com", "benfranklinplumbingaz.com",
    "serviceexperts.com", "rotorooter.com",
}

to_remove_rows = []
for row in all_rows:
    name = row[1].value.lower()
    if "solar" in name:
        to_remove_rows.append((row[0].value, row[1].value, "solar"))
    elif re.search(r"\bced\b|consolidated electrical distributor", name):
        to_remove_rows.append((row[0].value, row[1].value, "CED/distributor"))
    elif "strada services" in name:
        to_remove_rows.append((row[0].value, row[1].value, "Strada Services (large multi-state co, wrong target type)"))

remove_ids = {r[0] for r in to_remove_rows}
print(f"removing {len(remove_ids)} rows:")
for r in to_remove_rows:
    print(" ", r)

# ---- 4: same-domain grouping (on rows NOT being removed) ----
def clean_domain(website, domain):
    d = domain or website or ""
    d = re.sub(r"^https?://", "", str(d).strip())
    d = d.split("/")[0]
    if d.startswith("www."):
        d = d[4:]
    return d.lower()

by_domain = defaultdict(list)
for row in all_rows:
    if row[0].value in remove_ids:
        continue
    d = clean_domain(row[col["Website"]].value, row[col["Domain"]].value)
    if d and d not in ("facebook.com", "instagram.com"):
        by_domain[d].append(row)

flag_count_same_city = 0
flag_count_diff_city = 0
for domain, group in by_domain.items():
    if len(group) < 2 or domain in KNOWN_FRANCHISE_DOMAINS:
        continue
    cities = {(r[col["City"]].value or "").strip().lower() for r in group}
    other_ids = {r[0].value for r in group}
    for row in group:
        others = ", ".join(sorted(other_ids - {row[0].value}))
        if len(cities) == 1:
            note = f"Shares domain '{domain}' with {others} (SAME city) -- likely duplicate, verify before treating as separate targets"
            flag_count_same_city += 1
        else:
            note = f"Shares domain '{domain}' with {others} (different city) -- could be a multi-location franchise or a domain-scrape mismatch from sourcing, verify before outreach"
            flag_count_diff_city += 1
        existing_dc = row[col["Data Check"]].value or ""
        row[col["Data Check"]].value = f"{existing_dc}; {note}" if existing_dc else note

print(f"\nflagged {flag_count_same_city} rows as likely same-city duplicates")
print(f"flagged {flag_count_diff_city} rows as possible franchise/mismatch (different city)")

# capture Data Check values BEFORE delete_rows mutates the sheet (Cell objects would
# otherwise become stale references after rows shift)
local_flag_map = {row[0].value: row[col["Data Check"]].value for row in all_rows
                   if row[0].value not in remove_ids and row[col["Data Check"]].value}

# ---- apply removals (delete from bottom up to keep row indices valid) ----
rows_to_delete = sorted([row[0].row for row in all_rows if row[0].value in remove_ids], reverse=True)
for r in rows_to_delete:
    ws.delete_rows(r)

wb.save(MASTER)
print(f"\nlocal xlsx saved: {len(remove_ids)} rows removed, flags applied")

# ---- sync live Google Sheet: same removals + flags ----
gc = gspread.service_account(filename=str(Path.home() / "dev/hvac-lead-sourcing/service_account.json"))
sh = gc.open_by_key("1bbOBPow3M9a4dgEtQ2wodtdlkXy_fcAiyexKz7yfsQ0")
gsheet_ws = sh.worksheet("Enriched Master")
all_values = gsheet_ws.get_all_values()
gheader = all_values[0]
gcol = {name: i for i, name in enumerate(gheader)}

grows_to_delete = sorted([i + 2 for i, r in enumerate(all_values[1:]) if r and r[0] in remove_ids], reverse=True)
for r in grows_to_delete:
    gsheet_ws.delete_rows(r)
print(f"live Google Sheet: removed {len(grows_to_delete)} rows")

# reload for flag updates (row numbers shifted after deletes)
all_values2 = gsheet_ws.get_all_values()
id_to_row = {r[0]: i + 2 for i, r in enumerate(all_values2[1:]) if r and r[0]}
cell_updates = []
for rid, dc_value in local_flag_map.items():
    if rid in id_to_row:
        cell_updates.append(gspread.Cell(id_to_row[rid], gcol["Data Check"] + 1, dc_value))
if cell_updates:
    gsheet_ws.update_cells(cell_updates, value_input_option="RAW")
print(f"live Google Sheet: updated {len(cell_updates)} Data Check cells")

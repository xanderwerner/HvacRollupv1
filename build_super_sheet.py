#!/usr/bin/env python3
"""Merge master + ZoomInfo (free) + Apollo cross-check + ROC into one confidence-scored view.

Writes the 'Super Enrichment' Google Sheet: one row per company, "best" value per
field plus a confidence verdict (Verified / Single-source / Conflict), and explicit
franchise-parent flags so nobody dials a corporate HQ number by mistake.
"""
import json
import re
import time
from pathlib import Path

import gspread
import openpyxl

DATA = Path(__file__).parent / "data"
SHEET_ID = open("/private/tmp/claude-501/-Users-xanderwerner-Desktop/d86a8061-54d6-4adb-ac39-862af4658b11/scratchpad/super_sheet_id.txt").read().strip()


def load_jsonl(name, key="row_id"):
    d = {}
    p = DATA / name
    if not p.exists():
        return d
    for line in open(p):
        try:
            r = json.loads(line)
        except Exception:
            continue
        d[r[key]] = r
    return d


def namekey(n):
    return set(re.sub(r"[^a-z ]", "", str(n or "").lower()).split())


def names_agree(a, b):
    ka, kb = namekey(a), namekey(b)
    if not ka or not kb:
        return False
    # last-token heuristic isn't reliable with suffixes (Jr/III); use any shared token len>=3
    return any(len(t) >= 3 for t in ka & kb)


def main():
    wb = openpyxl.load_workbook(DATA / "AZ_targets_enriched_master.xlsx")
    master_rows = [r for r in list(wb["Enriched Master"].iter_rows(values_only=True))[1:] if r[1]]

    zi = load_jsonl("zi_enrich.jsonl")
    apv = load_jsonl("apollo_verify.jsonl")
    roc = load_jsonl("roc_names.jsonl")
    bad_roc_ids = set(json.load(open(DATA / "roc_bad_matches.json"))) if (DATA / "roc_bad_matches.json").exists() else set()
    roc = {k: v for k, v in roc.items() if k not in bad_roc_ids}

    hb = openpyxl.load_workbook(DATA / "ICP_hotlist_v2.xlsx")
    tier_by_id = {r[1]: str(r[0]).split(" -")[0].strip() for r in
                  list(hb["ICP Hot List v2"].iter_rows(values_only=True))[1:] if r[0] and r[1]}

    out_rows = []
    stats = {"owner_verified": 0, "owner_single": 0, "owner_conflict": 0, "owner_none": 0,
             "size_verified": 0, "size_single": 0, "size_none": 0, "franchise_flag": 0}

    for r in master_rows:
        (rid, name, trade, city, state, owner, title, cell, dnc, email, office,
         emp, rev, ebitda, founded, datacheck, cellsrc, foundvia, srcorig, coowners,
         li, website, domain, otherloc, lic, licclass, rating, reviews, notes) = r[:29]

        z = zi.get(rid)
        a = apv.get(rid)
        rc = roc.get(rid)

        is_franchise = bool(z and z.get("zi_state") and z["zi_state"] != "Arizona")
        z_usable = z and z.get("match") in ("domain", "name+AZ") and not is_franchise

        # ---- Owner best + confidence ----
        owner_best = owner
        sources_agree = []
        sources_conflict = []
        if a and a.get("matched") and a.get("apollo_name"):
            (sources_agree if names_agree(owner, a["apollo_name"]) else sources_conflict).append(("Apollo", a["apollo_name"]))
        if z_usable and z.get("zi_owner"):
            (sources_agree if names_agree(owner, z["zi_owner"]) else sources_conflict).append(("ZoomInfo", z["zi_owner"]))

        if not owner:
            if z_usable and z.get("zi_owner"):
                owner_best = z["zi_owner"]
                owner_conf = "ZI-only (unverified)"
                stats["owner_single"] += 1
            else:
                owner_conf = "NONE FOUND"
                stats["owner_none"] += 1
        elif sources_conflict:
            owner_conf = "CONFLICT: " + "; ".join(f"{s}={n}" for s, n in sources_conflict)
            stats["owner_conflict"] += 1
        elif sources_agree:
            owner_conf = "Verified (" + "+".join(s for s, _ in sources_agree) + ")"
            stats["owner_verified"] += 1
        else:
            owner_conf = "Single-source" + (f" ({foundvia})" if foundvia else "")
            stats["owner_single"] += 1

        # ---- Size best + confidence ----
        emp_best, rev_best = emp, rev
        size_notes = []
        try:
            emp_num = int(emp) if emp not in (None, "") else None
        except (TypeError, ValueError):
            emp_num = None
        if z_usable and z.get("zi_employees"):
            if emp_num:
                ratio = z["zi_employees"] / emp_num
                if 0.5 <= ratio <= 2.0:
                    size_notes.append(f"ZI agrees (~{z['zi_employees']})")
                    stats["size_verified"] += 1
                else:
                    size_notes.append(f"ZI CONFLICT: {z['zi_employees']} vs ours {emp}")
                    stats["size_verified"] += 1  # still counted as cross-checked
            else:
                emp_best = z["zi_employees"]
                size_notes.append("ZI-only")
                stats["size_single"] += 1
        elif emp:
            stats["size_single"] += 1
        else:
            stats["size_none"] += 1
        if z_usable and z.get("zi_revenue_k") and not rev_best:
            rev_best = z["zi_revenue_k"] * 1000
        if a and a.get("apollo_employees") and not emp_best:
            emp_best = a["apollo_employees"]
        if a and a.get("apollo_revenue") and not rev_best:
            rev_best = a["apollo_revenue"]

        flags = []
        if is_franchise:
            flags.append(f"FRANCHISE/PARENT MATCH in ZI ({z.get('zi_name')}, {z.get('zi_city')}, {z.get('zi_state')}) — its size/owner NOT used")
            stats["franchise_flag"] += 1
        if datacheck:
            flags.append(str(datacheck))
        if rc and rc.get("licenses") and not any(l.get("status") == "Active" for l in rc["licenses"]):
            flags.append("ROC license inactive")

        zi_owner_disp = f"{z.get('zi_owner','')} / {z.get('zi_owner_title','')}".strip(" /") if z else ""
        roc_owner_disp = ""
        if rc and rc.get("licenses"):
            active = [l for l in rc["licenses"] if l.get("status") == "Active"]
            lic0 = active[0] if active else rc["licenses"][0]
            roc_owner_disp = f"{lic0.get('qp','')} / {lic0.get('no','')}"

        out_rows.append([
            rid, name, trade, city, state, tier_by_id.get(rid, ""),
            owner_best or "", cell or "", dnc or "", cellsrc or "",
            emp_best or "", rev_best or "", ebitda or "",
            owner_conf, "; ".join(size_notes) or ("Single-source" if emp_best else "NONE"),
            "YES" if (z and z.get("zi_has_mobile")) else ("no" if z_usable else ""),
            zi_owner_disp,
            z.get("zi_employees", "") if z_usable else "",
            (z.get("zi_revenue_k", "") * 1000) if (z_usable and z.get("zi_revenue_k")) else "",
            z.get("zi_company_id", "") if z else "",
            a.get("apollo_name", "") if a else "",
            a.get("apollo_employees", "") if a else "",
            roc_owner_disp,
            office or "", website or "", li or (a.get("linkedin","") if a else ""), rating or "", reviews or "",
            "DIAL" if (owner_best and cell) else ("SKIP TRACE" if owner_best else "NAME NEEDED"),
            " | ".join(flags),
        ])

    print("merge stats:", stats)

    sa = gspread.service_account(filename=str(Path.home() / "dev/hvac-lead-sourcing/service_account.json"))
    for attempt in range(5):
        try:
            sh = sa.open_by_key(SHEET_ID)
            break
        except Exception as e:
            print("retry open", attempt, str(e)[:80]); time.sleep(8 * (attempt + 1))
    ws = sh.worksheet("Super Enrichment")
    ws.update(values=out_rows, range_name="A2", value_input_option="USER_ENTERED")
    print(f"pushed {len(out_rows)} rows to Super Enrichment")


if __name__ == "__main__":
    main()

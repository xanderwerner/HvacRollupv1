#!/usr/bin/env python3
"""Apply PE/ownership research findings to the master as Data Check notes (tagging only,
not removing -- whether to exclude PE-backed/corporate-owned rows is Xander's call)."""
import json
from pathlib import Path

import gspread
import openpyxl

DATA = Path(__file__).parent / "data"
MASTER = DATA / "AZ_targets_enriched_master.xlsx"

results = json.load(open(DATA / "pe_final_results.json"))
by_id = {r["id"]: r for r in results}

wb = openpyxl.load_workbook(MASTER)
ws = wb["Enriched Master"]
header = [c.value for c in ws[1]]
col = {n: i for i, n in enumerate(header)}

updated = []
for row in ws.iter_rows(min_row=2):
    if not row[1].value or row[0].value not in by_id:
        continue
    r = by_id[row[0].value]
    ot = r["ownership_type"]
    brand = r.get("parent_or_brand", "")
    note = r.get("note", "") or f"{r.get('source','')}"

    if ot == "PE-backed":
        tag = f"PE-BACKED ({brand}) -- {note} -- likely not a viable acquisition target via owner outreach"
    elif ot == "corporate-owned (not PE)":
        tag = f"CORPORATE-OWNED ({brand}) -- {note} -- not independently owned, likely not a viable target"
    elif ot == "independent franchisee":
        tag = f"Independent franchisee of {brand} (locally owned, franchisor may be PE-backed) -- {note}"
    elif ot == "unclear":
        tag = f"Ownership unclear after both free research and ZoomInfo check -- {note}"
    else:
        tag = ""  # independent/family-owned -- confirmed clean, no tag needed

    if tag:
        existing = row[col["Data Check"]].value or ""
        row[col["Data Check"]].value = f"{existing}; {tag}" if existing else tag
        updated.append((row[0].value, ot))

wb.save(MASTER)
print(f"local xlsx: tagged {len(updated)} rows")
from collections import Counter
print(Counter(u[1] for u in updated))

gc = gspread.service_account(filename=str(Path.home() / "dev/hvac-lead-sourcing/service_account.json"))
sh = gc.open_by_key("1bbOBPow3M9a4dgEtQ2wodtdlkXy_fcAiyexKz7yfsQ0")
gsheet_ws = sh.worksheet("Enriched Master")
all_values = gsheet_ws.get_all_values()
gheader = all_values[0]
gcol = {n: i for i, n in enumerate(gheader)}
id_to_row = {r[0]: i + 2 for i, r in enumerate(all_values[1:]) if r and r[0]}

wb2 = openpyxl.load_workbook(MASTER)
ws2 = wb2["Enriched Master"]
cell_updates = []
for row in ws2.iter_rows(min_row=2):
    if row[0].value in [u[0] for u in updated] and row[0].value in id_to_row:
        cell_updates.append(gspread.Cell(id_to_row[row[0].value], gcol["Data Check"] + 1, row[col["Data Check"]].value))
if cell_updates:
    gsheet_ws.update_cells(cell_updates, value_input_option="RAW")
print(f"live Google Sheet: updated {len(cell_updates)} Data Check cells")

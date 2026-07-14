#!/usr/bin/env python3
"""Clear 13 pre-existing '#ERROR!' values in the Owner Cell column (predates this
session -- an earlier Apollo phone-reveal attempt wrote the literal error string
instead of a real number or leaving it blank). A broken placeholder is worse than an
empty cell -- it looks like real coverage but isn't."""
from pathlib import Path

import gspread
import openpyxl

DATA = Path(__file__).parent / "data"
MASTER = DATA / "AZ_targets_enriched_master.xlsx"

wb = openpyxl.load_workbook(MASTER)
ws = wb["Enriched Master"]
header = [c.value for c in ws[1]]
col = {n: i for i, n in enumerate(header)}

fixed = []
for row in ws.iter_rows(min_row=2):
    if not row[1].value:
        continue
    if row[col["Owner Cell"]].value == "#ERROR!":
        row[col["Owner Cell"]].value = ""
        row[col["Cell Source"]].value = ""
        existing = row[col["Notes"]].value or ""
        note = ("Owner Cell previously showed a corrupted '#ERROR!' value from an earlier "
                "Apollo phone-reveal attempt (predates this session) -- cleared rather than "
                "left broken. Re-attempt via Apollo once direct-dial credits are available "
                "(exhausted until 2027-06-21) or ZoomInfo once a new credit cycle starts.")
        row[col["Notes"]].value = f"{existing}; {note}" if existing else note
        fixed.append(row[0].value)

wb.save(MASTER)
print(f"local xlsx: cleared {len(fixed)} broken Owner Cell values -- {fixed}")

gc = gspread.service_account(filename=str(Path.home() / "dev/hvac-lead-sourcing/service_account.json"))
sh = gc.open_by_key("1bbOBPow3M9a4dgEtQ2wodtdlkXy_fcAiyexKz7yfsQ0")
gsheet_ws = sh.worksheet("Enriched Master")
all_values = gsheet_ws.get_all_values()
gheader = all_values[0]
gcol = {n: i for i, n in enumerate(gheader)}
id_to_row = {r[0]: i + 2 for i, r in enumerate(all_values[1:]) if r and r[0]}

wb2 = openpyxl.load_workbook(MASTER)
ws2 = wb2["Enriched Master"]
cell_updates = []
for row in ws2.iter_rows(min_row=2):
    if row[0].value in fixed and row[0].value in id_to_row:
        grow = id_to_row[row[0].value]
        cell_updates.append(gspread.Cell(grow, gcol["Owner Cell"] + 1, ""))
        cell_updates.append(gspread.Cell(grow, gcol["Cell Source"] + 1, ""))
        cell_updates.append(gspread.Cell(grow, gcol["Notes"] + 1, row[col["Notes"]].value))
if cell_updates:
    gsheet_ws.update_cells(cell_updates, value_input_option="RAW")
print(f"live Google Sheet: updated {len(fixed)} rows ({len(cell_updates)} cells)")

#!/usr/bin/env python3
"""Write the 11 Apollo-confirmed (domain-verified) owner matches into both master artifacts.
Companies with 2 confirmed people get the primary in Owner Name/Title, the second in Co-Owners.
"""
import json
import re
from pathlib import Path

import gspread
import openpyxl

DATA = Path(__file__).parent / "data"
MASTER = DATA / "AZ_targets_enriched_master.xlsx"

confirmed = json.load(open(DATA / "az_apollo_confirmed_owners.json"))

# domain each confirmed contact's email belongs to, used to match to the right master row
by_domain = {}
for c in confirmed:
    dom = c["email"].split("@")[1].lower()
    by_domain.setdefault(dom, []).append(c)

wb = openpyxl.load_workbook(MASTER)
ws = wb["Enriched Master"]
header = [cell.value for cell in ws[1]]
col = {name: i for i, name in enumerate(header)}

def clean_domain(website, domain):
    if domain:
        return domain.strip().lower()
    if website:
        d = re.sub(r"^https?://", "", website.strip())
        d = d.split("/")[0]
        if d.startswith("www."):
            d = d[4:]
        return d.lower()
    return None

updated_rows = []
for row in ws.iter_rows(min_row=2):
    if not row[1].value:
        continue
    dom = clean_domain(row[col["Website"]].value, row[col["Domain"]].value)
    if dom not in by_domain:
        continue
    people = by_domain[dom]
    primary = people[0]
    row[col["Owner Name"]].value = primary["name"]
    row[col["Owner Title"]].value = primary["title"]
    row[col["Owner Email"]].value = primary["email"]
    row[col["Owner LinkedIn"]].value = primary["linkedin"]
    row[col["Owner Found Via"]].value = "Apollo (domain-verified)"
    if len(people) > 1:
        row[col["Co-Owners"]].value = "; ".join(f"{p['name']} ({p['title']})" for p in people[1:])
    updated_rows.append((row[0].value, row[1].value, primary["name"]))

wb.save(MASTER)
print(f"local xlsx: updated {len(updated_rows)} rows")
for r in updated_rows:
    print(" ", r)

gc = gspread.service_account(filename=str(Path.home() / "dev/hvac-lead-sourcing/service_account.json"))
sh = gc.open_by_key("1bbOBPow3M9a4dgEtQ2wodtdlkXy_fcAiyexKz7yfsQ0")
gsheet_ws = sh.worksheet("Enriched Master")
all_values = gsheet_ws.get_all_values()
gheader = all_values[0]
gcol = {name: i for i, name in enumerate(gheader)}
id_to_row = {r[0]: i + 2 for i, r in enumerate(all_values[1:]) if r and r[0]}

cell_updates = []
for rid, name, owner_name in updated_rows:
    if rid not in id_to_row:
        continue
    r = id_to_row[rid]
    people = by_domain[[d for d, ps in by_domain.items() if any(p["name"] == owner_name for p in ps)][0]]
    primary = people[0]
    cell_updates.append(gspread.Cell(r, gcol["Owner Name"] + 1, primary["name"]))
    cell_updates.append(gspread.Cell(r, gcol["Owner Title"] + 1, primary["title"]))
    cell_updates.append(gspread.Cell(r, gcol["Owner Email"] + 1, primary["email"]))
    cell_updates.append(gspread.Cell(r, gcol["Owner LinkedIn"] + 1, primary["linkedin"]))
    cell_updates.append(gspread.Cell(r, gcol["Owner Found Via"] + 1, "Apollo (domain-verified)"))
    if len(people) > 1:
        cell_updates.append(gspread.Cell(r, gcol["Co-Owners"] + 1, "; ".join(f"{p['name']} ({p['title']})" for p in people[1:])))

gsheet_ws.update_cells(cell_updates, value_input_option="RAW")
print(f"live Google Sheet: updated {len(updated_rows)} rows ({len(cell_updates)} cells)")

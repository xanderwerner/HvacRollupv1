#!/usr/bin/env python3
"""Restore the '#ERROR!' marker in Owner Cell for the same 13 rows, as an intentional
visual flag for Xander's buddy (using his own ZoomInfo trial account) to spot and
re-look-up -- not a data error this time, a deliberate to-do marker."""
from pathlib import Path

import gspread
import openpyxl

DATA = Path(__file__).parent / "data"
MASTER = DATA / "AZ_targets_enriched_master.xlsx"

TARGET_IDS = {'C028', 'C042', 'C050', 'C051', 'C064', 'C112', 'C115', 'C148',
              'C154', 'C157', 'C185', 'C632', 'C643'}

wb = openpyxl.load_workbook(MASTER)
ws = wb["Enriched Master"]
header = [c.value for c in ws[1]]
col = {n: i for i, n in enumerate(header)}

restored = []
for row in ws.iter_rows(min_row=2):
    if row[0].value in TARGET_IDS:
        row[col["Owner Cell"]].value = "#ERROR!"
        row[col["Cell Source"]].value = "NEEDS LOOKUP"
        restored.append(row[0].value)

wb.save(MASTER)
print(f"local xlsx: restored #ERROR! flag on {len(restored)} rows -- {restored}")

gc = gspread.service_account(filename=str(Path.home() / "dev/hvac-lead-sourcing/service_account.json"))
sh = gc.open_by_key("1bbOBPow3M9a4dgEtQ2wodtdlkXy_fcAiyexKz7yfsQ0")
gsheet_ws = sh.worksheet("Enriched Master")
all_values = gsheet_ws.get_all_values()
gheader = all_values[0]
gcol = {n: i for i, n in enumerate(gheader)}
id_to_row = {r[0]: i + 2 for i, r in enumerate(all_values[1:]) if r and r[0]}

cell_updates = []
for rid in restored:
    if rid in id_to_row:
        grow = id_to_row[rid]
        cell_updates.append(gspread.Cell(grow, gcol["Owner Cell"] + 1, "#ERROR!"))
        cell_updates.append(gspread.Cell(grow, gcol["Cell Source"] + 1, "NEEDS LOOKUP"))
if cell_updates:
    gsheet_ws.update_cells(cell_updates, value_input_option="RAW")
print(f"live Google Sheet: updated {len(restored)} rows ({len(cell_updates)} cells)")
